from openpyxl import load_workbook
import pandas as pd
import numpy as np
class Extract:
    def __init__(self, path,base, flag):
        self.path = path
        self.pathbase = base
        self.flag = flag
        self.tableNum = 64
        self.tableRow = 21
        self.tableLine = 21
        self.wb = load_workbook(self.path)
        print(f"路径为: {self.pathbase}")
        self.wb_base = load_workbook(self.pathbase)
        self.ws_base = self.wb_base.active
        self.num_rowsbase = self.ws_base.max_row
        self.num_colsbase = self.ws_base.max_column
        self.ws = self.wb.active
        self.num_rows = self.ws.max_row
        self.num_cols = self.ws.max_column

    def transformZero(self,ws,is_base):
        w,h = 0,0
        if is_base:
            w = ws.max_row
            h = ws.max_column
        else:
            w = ws.max_row
            h = ws.max_column-2
        for i in range(3, w+1):
            for j in range(2, h+1):
                # 如果单元格的值为"OVER"
                if ws.cell(row=i, column=j).value == 'OVER':
                    if i>2 and i < w-2 and j>1 and j < h-2:
                        if ws.cell(row=i-1, column=j).value != 'OVER':
                            ws.cell(row=i-1, column=j).value = 0
                        if ws.cell(row=i+1, column=j).value != 'OVER':
                            ws.cell(row=i+1, column=j).value = 0
                        if ws.cell(row=i, column=j-1).value != 'OVER':
                            ws.cell(row=i, column=j-1).value = 0
                        if ws.cell(row=i+1, column=j+1).value != 'OVER':
                            ws.cell(row=i+1, column=j+1).value = 0


        for i in range(3, w+1):
            for j in range(2, h+1):
                # 如果单元格的值为"OVER"
                # print(i,j)
                # print(ws.cell(row=i, column=j).value)
                if ws.cell(row=i, column=j).value == 'OVER' or ws.cell(row=i, column=j).value < 0:
                    ws.cell(row=i, column=j).value = 0

    def subbias(self, ws,ws_base):
        i=3
        n = 0
        while(i<self.num_rows+1):
            j=2
            while(j<self.num_colsbase+1):
                if(ws.cell(row=i,column=j).value!=0):
                    # print(ws.cell(row=i,column=j).value,ws_base.cell(row=n+3,column=j).value)
                    # print("loc is ",n+3,j,i,j)
                    ws.cell(row=i,column=j).value=ws.cell(row=i,column=j).value-ws_base.cell(row=n+3,column=j).value
                    if(ws.cell(row=i,column=j).value<0):
                        ws.cell(row=i,column=j).value=0
                j+=1
            n+=1
            i+=1
            if n==20:
                i+=1
                n =0


    # 提取每个表的数据
    def every(self):
        self.transformZero(self.ws_base,True)  # 先调用transformZero方法处理数据
        # for row in self.ws_base.iter_rows(values_only=True):
        #     print(row)
        self.transformZero(self.ws,False)  # 先调用transformZero方法处理数据
        self.subbias(self.ws,self.ws_base)
        lable = []  # 保存表名
        count_data = []  # 保存提取的数据
        num = []  # 保存所有数据，用于归一化
        i = 2
        max_length_column_23 = 0
        max_length_column_24 = 0
        for row in range(2, self.ws.max_row + 1,22):
            if self.ws.cell(row=row, column=23).value:
                max_length_column_23 = max(max_length_column_23, len(self.ws.cell(row=row, column=23).value))
        
        for row in range(2, self.ws.max_row + 1,22):
            if self.ws.cell(row=row, column=24).value:
                max_length_column_24 = max(max_length_column_24, len(self.ws.cell(row=row, column=24).value))
        #max_length_column_23 = max(len(self.ws.cell(row=row, column=23).value) for row in range(2, self.ws.max_row + 1,22))
        #max_length_column_24 = max(len(self.ws.cell(row=row, column=24).value) for row in range(2, self.ws.max_row + 1,22))
        while i < self.num_rows:
            table_start_row = i  # 确定每个表的起始行
            # 获取表名
            #print(table_start_row)

            #print(self.ws.cell(row=table_start_row, column=23).value ,self.ws.cell(row=table_start_row,
            #                                                                              column=24).value)
            
            data_head = self.ws.cell(row=table_start_row, column=23).value.replace(" ", "N") if self.ws.cell(row=table_start_row, column=23).value else ""
            data_tail = self.ws.cell(row=table_start_row, column=24).value.replace(" ", "N") if self.ws.cell(row=table_start_row, column=24).value else ""
            if len(data_head) < max_length_column_23 :
                str = "N" * (max_length_column_23 -len(data_head))
                data_head = data_head + str
            if len(data_tail) < max_length_column_24 :
                str = "N" * (max_length_column_24 -len(data_tail))
                data_tail = data_tail + str
            tableName = data_head + '-' + data_tail
            # tableName = self.ws.cell(row=table_start_row, column=23).value + '-' + self.ws.cell(row=table_start_row,
            #                                                                               column=24).value
            data = []  # 保存当前表的数据
            temp_data = []  # 临时数据列表
            for row_idx in range(i + 1,  i + self.tableLine + 1):
                # 提取每一行的数据
                row_data = [cell[0].value for cell in
                            self.ws.iter_cols(min_row=row_idx, max_row=row_idx, min_col=2, max_col=self.tableRow + 1)]
                # print('row_data',row_data)
                # print("row_data_shape",type(row_data))
                data.append(row_data)
            lable.append(tableName)  # 保存表名
            for row in data:
                for DNA in row:
                    temp_data.append(DNA)  # 将数据添加到临时列表
                    num.append(DNA)  # 将数据添加到总数据列表
            count_data.append(temp_data)  # 将临时数据添加到总数据列表
            i += 22
        # print(len(count_data[0]))  # 打印第一个表的数据长度
        #print(count_data, lable)
        max_num = max(num)  # 获取数据的最大值
        min_num = min(num)  # 获取数据的最小值
        # print(max_num)
        # print(min_num)
        # 对数据进行归一化处理
        #count = [[(x - min_num) / (max_num - min_num) for x in sublist] for sublist in count_data]
        count = [[np.log1p(x) for x in sublist] for sublist in count_data]
        count_data = count
        reshaped_data = []
        np.random.seed(42)  # 你可以选择任意的种子值

        for sublist in count_data:
            reshaped = np.array(sublist).reshape(21, 21)
            reshaped_data.append(reshaped)
        print("reshaped_data",len(reshaped_data))
        if self.flag == 1:
        # 当flag为True时，随机选择2000个数据
            #return reshaped_data, lable, max_num, min_num  # 返回数据和表名
            indices = np.random.choice(len(reshaped_data), 3500, replace=False)
            selected_data = [reshaped_data[i] for i in indices]
            selected_labels = [lable[i] for i in indices]
            return selected_data, selected_labels, max_num, min_num
        elif self.flag == 2:
            # 当flag为False时，返回剩余的数据
            indices = np.random.choice(len(reshaped_data), 3500, replace=False)
            remaining_indices = [i for i in range(len(reshaped_data)) if i not in indices]
            remaining_data = [reshaped_data[i] for i in remaining_indices]
            remaining_labels = [lable[i] for i in remaining_indices]
            return remaining_data[50:],remaining_labels[50:], max_num, min_num
        elif self.flag == 3:
            # 当flag为False时，返回剩余的数据
            indices = np.random.choice(len(reshaped_data), 3500, replace=False)
            remaining_indices = [i for i in range(len(reshaped_data)) if i not in indices]
            remaining_data = [reshaped_data[i] for i in remaining_indices]
            remaining_labels = [lable[i] for i in remaining_indices]
            return remaining_data[:50], remaining_labels[:50], max_num, min_num
        else:
            return reshaped_data, lable, max_num, min_num


import pandas as pd
import gensim
from gensim.models import Word2Vec

class DNAWord2Vec:
    def __init__(self, sequences):
        self.sequences = sequences

    def preprocess(self):
        # 将每个DNA序列转换为单词列表
        word_lists = []
        for seq in self.sequences:
            word_lists.append(list(seq))  # 每个碱基作为一个单词
        return word_lists

    def train_model(self, word_lists):
        # 训练Word2Vec模型
        model = Word2Vec(sentences=word_lists, vector_size=100, window=3, min_count=1, sg=1)  # sg=1表示使用skip-gram
        return model

# 主函数
if __name__ == '__main__':
    count, label, _, _ = Extract("./AgNCs/combined_data.xlsx","./AgNCs/bias.xlsx", 4).every()  # 创建Extract对象并调用every方法
    print('count_shape',count[0].shape)
    print("count_type",type(count))
    # print("label",label[0])
    # for i in range(20):
    #     print(label[i],len(label[i]))


    sentences = [list(seq) for seq in label]

    dna_model = DNAWord2Vec(sentences)
    word_lists = dna_model.preprocess()  # 数据预处理
    model = dna_model.train_model(word_lists)  # 训练Word2Vec模型

    # 保存模型
    model.save("dna_word2vec.model")

    # 使用模型
    vector_a = model.wv['A']  # 获取碱基A的向量
    print("Vector for A:", vector_a)



    # df=pd.DataFrame({'caption':label,'count':count})
    # names = [f"name_{i}" for i in np.arange(len(label))]
    # df.insert(0,"image",names)
    # df
    # df.to_csv('./data.csv')




    # a = Extract("./AgNCs/combined_data.xlsx","./AgNCs/bias.xlsx", 1).every()
    # print(a[0])
    # print(len(a[0]))
    # print(np.array(a[0]).shape)
